from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def export_table_to_excel(path, title, headers, rows):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Report"

    sheet["A1"] = title
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    sheet.append([])
    sheet.append(list(headers))

    for header_cell in sheet[3]:
        header_cell.font = Font(bold=True)
        header_cell.fill = PatternFill("solid", fgColor="D9E2F3")

    for row in rows:
        sheet.append(list(row))

    for column_index in range(1, len(headers) + 1):
        max_length = 0
        for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=column_index, max_col=column_index):
            cell = row[0]
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 2, 40)

    workbook.save(path)


def export_table_to_pdf(path, title, headers, rows):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(str(path), pagesize=letter, leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    story = [Paragraph(title, styles["Title"]), Spacer(1, 12)]

    table_data = [list(headers)] + [list(row) for row in rows]
    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
                ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
            ]
        )
    )
    story.append(table)
    doc.build(story)
